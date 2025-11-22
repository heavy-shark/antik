"""
Browser Profile Manager
Handles creation, deletion, and management of browser profiles
"""
import os
import json
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


class ProfileManager:
    def __init__(self):
        # Use profiles directory relative to app location
        app_dir = Path(__file__).parent
        self.profiles_dir = app_dir / "profiles"
        self.profiles_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_file = self.profiles_dir / "profiles_metadata.json"
        self.load_metadata()

    def load_metadata(self):
        """Load profiles metadata from JSON file"""
        if self.metadata_file.exists():
            with open(self.metadata_file, 'r', encoding='utf-8') as f:
                self.metadata = json.load(f)
        else:
            self.metadata = {}
            self.save_metadata()

    def save_metadata(self):
        """Save profiles metadata to JSON file"""
        with open(self.metadata_file, 'w', encoding='utf-8') as f:
            json.dump(self.metadata, f, indent=2)

    def create_profile(self, profile_name, description="", email="", password="", proxy="", twofa_secret=""):
        """Create a new browser profile with optional credentials"""
        if profile_name in self.metadata:
            return False, "Profile already exists"

        profile_path = self.profiles_dir / profile_name
        profile_path.mkdir(parents=True, exist_ok=True)

        self.metadata[profile_name] = {
            "name": profile_name,
            "description": description,
            "email": email,
            "password": password,
            "proxy": proxy,
            "twofa_secret": twofa_secret,
            "created_at": datetime.now().isoformat(),
            "last_used": None,
            "path": str(profile_path)
        }
        self.save_metadata()
        return True, "Profile created successfully"

    def delete_profile(self, profile_name):
        """Delete a browser profile"""
        if profile_name not in self.metadata:
            return False, "Profile does not exist"

        profile_path = Path(self.metadata[profile_name]["path"])
        if profile_path.exists():
            shutil.rmtree(profile_path)

        del self.metadata[profile_name]
        self.save_metadata()
        return True, "Profile deleted successfully"

    def get_all_profiles(self):
        """Get list of all profiles"""
        return list(self.metadata.keys())

    def get_profile_info(self, profile_name):
        """Get detailed info about a profile"""
        return self.metadata.get(profile_name, None)

    def update_last_used(self, profile_name):
        """Update last used timestamp for a profile"""
        if profile_name in self.metadata:
            self.metadata[profile_name]["last_used"] = datetime.now().isoformat()
            self.save_metadata()

    def get_profile_path(self, profile_name):
        """Get the filesystem path for a profile"""
        if profile_name in self.metadata:
            return self.metadata[profile_name]["path"]
        return None

    def import_from_excel(self, excel_path):
        """
        Import profiles from Excel file

        Excel format:
        Row 1 (headers): email | password | proxy | 2fa_secret
        Row 2+: actual data

        Args:
            excel_path: Path to Excel file

        Returns:
            tuple: (success_count, skipped_count, error_list)
        """
        try:
            # Load Excel file
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active

            success_count = 0
            skipped_count = 0
            errors = []

            # Start from row 2 (skip headers in row 1)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue

                    # Extract values from columns A-D
                    email = str(row[0]).strip() if row[0] else ""
                    password = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    proxy = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    twofa_secret = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    # Skip if no email (required field)
                    if not email or email == "None":
                        skipped_count += 1
                        errors.append(f"Row {row_num}: Missing email, skipped")
                        continue

                    # Use email as profile name (sanitize for filesystem)
                    profile_name = email.replace("@", "_at_").replace(".", "_")

                    # Check if profile already exists
                    if profile_name in self.metadata:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: Profile '{email}' already exists, skipped")
                        continue

                    # Create profile
                    success, message = self.create_profile(
                        profile_name=profile_name,
                        description=f"Imported from Excel (Row {row_num})",
                        email=email,
                        password=password,
                        proxy=proxy,
                        twofa_secret=twofa_secret
                    )

                    if success:
                        success_count += 1
                    else:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {message}")

                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Row {row_num}: {str(e)}")

            workbook.close()
            return success_count, skipped_count, errors

        except Exception as e:
            return 0, 0, [f"Failed to read Excel file: {str(e)}"]
