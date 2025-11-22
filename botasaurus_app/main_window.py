"""
Hysk Mexc Futures - Main Window
Dark blue theme UI inspired by NaVI Blue design
"""
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QTableWidget,
    QTableWidgetItem, QHeaderView, QCheckBox, QRadioButton,
    QButtonGroup, QMessageBox, QFileDialog, QFrame
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QFont, QColor
from profile_manager import ProfileManager
from scraper_runner import ScraperRunner, ScraperThread, CheckProxyThread, ManualBrowserThread, MexcLoginThread, MexcShortThread, MexcLongThread
import json
import pyotp
import time


class MainWindow(QMainWindow):
    """Main application window with dark blue theme"""

    def __init__(self):
        super().__init__()
        self.profile_manager = ProfileManager()
        self.scraper_runner = ScraperRunner(self.profile_manager)
        self.version = "v0.3"
        self.totp_data = {}  # Store TOTP data for each row: {row: secret}
        self.active_threads = {}  # Store active threads (login) by profile name - PARALLEL execution
        self.active_drivers = {}  # Store active Driver instances to prevent garbage collection
        self.active_browser_threads = {}  # Store ManualBrowserThread instances with exec()
        self.active_trade_threads = {}  # Store ShortLongTradeThread instances

        # Single global timer for all TOTP updates (performance optimization)
        self.global_totp_timer = QTimer()
        self.global_totp_timer.timeout.connect(self.update_all_totp_codes)

        self.init_ui()

    def init_ui(self):
        """Initialize the user interface with dark blue theme"""
        self.setWindowTitle(f"Hysk Mexc Futures {self.version}")
        self.setGeometry(100, 100, 1400, 900)

        # Apply dark blue theme
        self.setStyleSheet("""
            QMainWindow {
                background-color: #0a1929;
            }
            QWidget {
                background-color: #0a1929;
                color: #ffffff;
                font-family: 'Segoe UI', Arial;
                font-size: 12px;
            }
            QLabel {
                color: #ffffff;
            }
            QPushButton {
                background-color: #1976d2;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QPushButton:pressed {
                background-color: #0d47a1;
            }
            QPushButton#settingsBtn {
                background-color: #1e3a5f;
                padding: 8px 16px;
            }
            QPushButton#settingsBtn:hover {
                background-color: #2d4a6f;
            }
            QPushButton#importBtn {
                background-color: #1976d2;
            }
            QPushButton#openSelectedBtn {
                background-color: #1976d2;
            }
            QPushButton#closeSelectedBtn {
                background-color: #455a64;
            }
            QPushButton#closeSelectedBtn:hover {
                background-color: #546e7a;
            }
            QPushButton#clearLogBtn {
                background-color: #1e3a5f;
                padding: 6px 12px;
                font-size: 11px;
            }
            QTableWidget {
                background-color: #0d1b2a;
                border: 1px solid #1e3a5f;
                gridline-color: #1e3a5f;
                color: #ffffff;
            }
            QTableWidget QWidget {
                background-color: #0d1b2a;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #1e3a5f;
                height: 40px;
            }
            QTableWidget::item:selected {
                background-color: #1e3a5f;
            }
            QHeaderView::section {
                background-color: #0d1b2a;
                color: #64b5f6;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #1e3a5f;
                border-right: 1px solid #1e3a5f;
                font-weight: bold;
            }
            QTextEdit {
                background-color: #0d1b2a;
                border: 1px solid #1e3a5f;
                color: #4caf50;
                font-family: 'Consolas', 'Courier New', monospace;
                padding: 8px;
            }
            QRadioButton {
                color: #ffffff;
                spacing: 8px;
            }
            QRadioButton::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #1976d2;
                border-radius: 10px;
                background-color: #0d1b2a;
            }
            QRadioButton::indicator:checked {
                background-color: #1976d2;
                border: 2px solid #64b5f6;
            }
            QRadioButton::indicator:hover {
                border: 2px solid #64b5f6;
            }
            QCheckBox {
                spacing: 5px;
                background: none;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border: 2px solid #1976d2;
                border-radius: 8px;
                background-color: rgba(0, 0, 0, 0);
            }
            QCheckBox::indicator:unchecked {
                background-color: rgba(0, 0, 0, 0);
                background: none;
            }
            QCheckBox::indicator:checked {
                background-color: #1976d2;
                border: 2px solid #1976d2;
            }
            QCheckBox::indicator:unchecked:hover {
                background-color: rgba(0, 0, 0, 0);
                border: 2px solid #64b5f6;
            }
            QCheckBox::indicator:checked:hover {
                background-color: #1976d2;
                border: 2px solid #64b5f6;
            }
            QFrame#separator {
                background-color: #1e3a5f;
            }
        """)

        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # === HEADER ===
        header_widget = self.create_header()
        main_layout.addWidget(header_widget)

        # Content area with padding
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(20)

        # === PROFILES SECTION ===
        profiles_section = self.create_profiles_section()
        content_layout.addWidget(profiles_section)

        # === OPERATION MODE SECTION ===
        operation_section = self.create_operation_mode_section()
        content_layout.addWidget(operation_section)

        # === SHORT/LONG SETTINGS SECTION ===
        self.short_long_section = self.create_short_long_settings_section()
        content_layout.addWidget(self.short_long_section)
        self.short_long_section.hide()  # Hidden by default

        # === LOGS SECTION ===
        logs_section = self.create_logs_section()
        content_layout.addWidget(logs_section)

        main_layout.addWidget(content_widget)

        # Load profiles into table
        self.refresh_profiles_table()

    def create_header(self):
        """Create header with title and settings button"""
        header = QWidget()
        header.setFixedHeight(60)
        header.setStyleSheet("background-color: #0d1b2a; border-bottom: 2px solid #1e3a5f;")

        layout = QHBoxLayout(header)
        layout.setContentsMargins(20, 0, 20, 0)

        # Title
        title = QLabel(f"Hysk Mexc Futures {self.version}")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")

        layout.addWidget(title)
        layout.addStretch()

        # Delete Selected button (hidden by default)
        self.delete_selected_btn = QPushButton("🗑️ Delete Selected")
        self.delete_selected_btn.setObjectName("deleteSelectedBtn")
        self.delete_selected_btn.setStyleSheet("background-color: #d32f2f; padding: 8px 16px;")
        self.delete_selected_btn.clicked.connect(self.delete_selected_profiles)
        self.delete_selected_btn.setFixedWidth(150)
        self.delete_selected_btn.hide()  # Hidden by default

        layout.addWidget(self.delete_selected_btn)

        # Settings button
        settings_btn = QPushButton("⚙ Settings")
        settings_btn.setObjectName("settingsBtn")
        settings_btn.clicked.connect(self.open_settings)
        settings_btn.setFixedWidth(120)

        layout.addWidget(settings_btn)

        return header

    def create_profiles_section(self):
        """Create profiles table section"""
        section = QWidget()
        layout = QVBoxLayout(section)
        layout.setSpacing(15)

        # Section title
        title = QLabel("Profiles")
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")
        layout.addWidget(title)

        # Profiles table
        self.profiles_table = QTableWidget()
        self.profiles_table.setColumnCount(5)
        self.profiles_table.setHorizontalHeaderLabels([
            "", "Email", "Proxy", "2FA Code", "Status"
        ])

        # Set column widths
        header = self.profiles_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # Select
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # Email
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # Proxy
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # 2FA Code
        header.setSectionResizeMode(4, QHeaderView.Fixed)  # Status

        self.profiles_table.setColumnWidth(0, 50)  # Select (smaller)
        self.profiles_table.setColumnWidth(3, 150)  # 2FA Code
        self.profiles_table.setColumnWidth(4, 100)  # Status

        self.profiles_table.setMinimumHeight(250)
        self.profiles_table.setMaximumHeight(350)

        layout.addWidget(self.profiles_table)

        # Buttons row
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        # Import button
        import_btn = QPushButton("📥 Import")
        import_btn.setObjectName("importBtn")
        import_btn.clicked.connect(self.import_profiles)
        import_btn.setFixedWidth(150)
        buttons_layout.addWidget(import_btn)

        buttons_layout.addStretch()

        # Open Selected button
        open_selected_btn = QPushButton("Open Selected")
        open_selected_btn.setObjectName("openSelectedBtn")
        open_selected_btn.clicked.connect(self.open_selected_profiles)
        open_selected_btn.setFixedWidth(150)
        buttons_layout.addWidget(open_selected_btn)

        # Close Selected button
        close_selected_btn = QPushButton("Close Selected")
        close_selected_btn.setObjectName("closeSelectedBtn")
        close_selected_btn.clicked.connect(self.close_selected_profiles)
        close_selected_btn.setFixedWidth(150)
        buttons_layout.addWidget(close_selected_btn)

        layout.addLayout(buttons_layout)

        return section

    def create_operation_mode_section(self):
        """Create operation mode section with radio buttons"""
        section = QWidget()
        layout = QVBoxLayout(section)
        layout.setSpacing(12)

        # Section title
        title = QLabel("Operation Mode")
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")
        layout.addWidget(title)

        # Radio buttons layout
        radio_layout = QHBoxLayout()
        radio_layout.setSpacing(30)

        # Create button group
        self.operation_group = QButtonGroup()

        # Create radio buttons
        modes = [
            ("Manual", "manual"),
            ("Login", "login"),
            ("Short", "short"),
            ("Long", "long"),
            ("Balance", "balance"),
            ("RK", "rk")
        ]

        for i, (label, value) in enumerate(modes):
            radio = QRadioButton(label)
            radio.setProperty("mode", value)
            self.operation_group.addButton(radio, i)
            radio_layout.addWidget(radio)

            # Select first option by default
            if i == 0:
                radio.setChecked(True)

        radio_layout.addStretch()
        layout.addLayout(radio_layout)

        # Connect signal to update Short/Long settings visibility
        self.operation_group.buttonClicked.connect(self.on_operation_mode_changed)

        return section

    def create_short_long_settings_section(self):
        """Create Short/Long mode settings section"""
        from PySide6.QtWidgets import QLineEdit, QComboBox, QRadioButton

        section = QWidget()
        layout = QVBoxLayout(section)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 15, 10, 15)

        # Section title
        title = QLabel("Trading Settings")
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")
        layout.addWidget(title)

        # === TOKEN LINK ===
        token_layout = QHBoxLayout()
        token_label = QLabel("Token Link:")
        token_label.setFixedWidth(120)
        token_label.setStyleSheet("color: #90caf9;")
        self.token_link_input = QLineEdit()
        self.token_link_input.setPlaceholderText("Enter token contract address or link...")
        self.token_link_input.setStyleSheet("""
            QLineEdit {
                background-color: #132f4c;
                border: 1px solid #1e4976;
                border-radius: 4px;
                padding: 8px;
                color: #ffffff;
            }
            QLineEdit:focus {
                border: 1px solid: #2196f3;
            }
        """)
        token_layout.addWidget(token_label)
        token_layout.addWidget(self.token_link_input)
        layout.addLayout(token_layout)

        # === POSITION IN % ===
        position_layout = QHBoxLayout()
        position_label = QLabel("Position in %:")
        position_label.setFixedWidth(120)
        position_label.setStyleSheet("color: #90caf9;")
        self.position_dropdown = QComboBox()
        self.position_dropdown.addItems(["25%", "50%", "75%", "100%", "Custom"])
        self.position_dropdown.setStyleSheet("""
            QComboBox {
                background-color: #132f4c;
                border: 1px solid #1e4976;
                border-radius: 4px;
                padding: 8px;
                color: #ffffff;
                min-width: 120px;
            }
            QComboBox:hover {
                border: 1px solid #2196f3;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #90caf9;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #132f4c;
                border: 1px solid #1e4976;
                selection-background-color: #1976d2;
                color: #ffffff;
            }
        """)
        position_layout.addWidget(position_label)
        position_layout.addWidget(self.position_dropdown)

        # Custom % input (initially hidden)
        self.custom_position_input = QLineEdit()
        self.custom_position_input.setPlaceholderText("Enter custom %...")
        self.custom_position_input.setFixedWidth(150)
        self.custom_position_input.setStyleSheet("""
            QLineEdit {
                background-color: #132f4c;
                border: 1px solid #1e4976;
                border-radius: 4px;
                padding: 8px;
                color: #ffffff;
            }
            QLineEdit:focus {
                border: 1px solid #2196f3;
            }
        """)
        self.custom_position_input.hide()  # Hidden by default
        position_layout.addWidget(self.custom_position_input)

        position_layout.addStretch()
        layout.addLayout(position_layout)

        # Connect dropdown change to show/hide custom input
        self.position_dropdown.currentTextChanged.connect(self.on_position_dropdown_changed)

        # === TYPE OF ZALIV (MARKET/LIMIT) ===
        zaliv_layout = QHBoxLayout()
        zaliv_label = QLabel("Type of Zaliv:")
        zaliv_label.setFixedWidth(120)
        zaliv_label.setStyleSheet("color: #90caf9;")
        zaliv_layout.addWidget(zaliv_label)

        # Radio buttons for Market/Limit
        self.zaliv_group = QButtonGroup()

        self.market_radio = QRadioButton("Market")
        self.market_radio.setChecked(True)  # Default
        self.market_radio.setStyleSheet("QRadioButton { color: #ffffff; }")
        self.zaliv_group.addButton(self.market_radio, 0)
        zaliv_layout.addWidget(self.market_radio)

        self.limit_radio = QRadioButton("Limit")
        self.limit_radio.setStyleSheet("QRadioButton { color: #ffffff; }")
        self.zaliv_group.addButton(self.limit_radio, 1)
        zaliv_layout.addWidget(self.limit_radio)

        zaliv_layout.addSpacing(20)

        # Limit price input (initially hidden)
        limit_price_label = QLabel("Limit Price:")
        limit_price_label.setStyleSheet("color: #90caf9;")
        self.limit_price_label = limit_price_label
        self.limit_price_label.hide()  # Hidden by default

        self.limit_price_input = QLineEdit()
        self.limit_price_input.setPlaceholderText("Enter limit price...")
        self.limit_price_input.setFixedWidth(180)
        self.limit_price_input.setStyleSheet("""
            QLineEdit {
                background-color: #132f4c;
                border: 1px solid #1e4976;
                border-radius: 4px;
                padding: 8px;
                color: #ffffff;
            }
            QLineEdit:focus {
                border: 1px solid #2196f3;
            }
        """)
        self.limit_price_input.hide()  # Hidden by default

        zaliv_layout.addWidget(self.limit_price_label)
        zaliv_layout.addWidget(self.limit_price_input)
        zaliv_layout.addStretch()
        layout.addLayout(zaliv_layout)

        # Connect radio button change to show/hide limit price
        self.zaliv_group.buttonClicked.connect(self.on_zaliv_type_changed)

        return section

    def on_operation_mode_changed(self, button):
        """Handle operation mode change - show/hide Short/Long settings"""
        mode = button.property("mode")

        # Show settings only for Short or Long modes
        if mode in ["short", "long"]:
            self.short_long_section.show()
            # Update title based on mode
            mode_name = "Short" if mode == "short" else "Long"
            # Find and update title (first child QLabel)
            for child in self.short_long_section.findChildren(QLabel):
                if child.font().bold():
                    child.setText(f"{mode_name} Trading Settings")
                    break
        else:
            self.short_long_section.hide()

    def on_position_dropdown_changed(self, text):
        """Show/hide custom position input based on dropdown selection"""
        if text == "Custom":
            self.custom_position_input.show()
            self.custom_position_input.setFocus()
        else:
            self.custom_position_input.hide()
            self.custom_position_input.clear()

    def on_zaliv_type_changed(self, button):
        """Show/hide limit price input based on Market/Limit selection"""
        if button == self.limit_radio:
            self.limit_price_label.show()
            self.limit_price_input.show()
            self.limit_price_input.setFocus()
        else:
            self.limit_price_label.hide()
            self.limit_price_input.hide()
            self.limit_price_input.clear()

    def get_short_long_settings(self):
        """Get current Short/Long trading settings"""
        # Get token link
        token_link = self.token_link_input.text().strip()

        # Get position percentage
        position_dropdown_value = self.position_dropdown.currentText()
        if position_dropdown_value == "Custom":
            position_percent = self.custom_position_input.text().strip()
        else:
            position_percent = position_dropdown_value.replace("%", "")

        # Get zaliv type
        zaliv_type = "Market" if self.market_radio.isChecked() else "Limit"

        # Get limit price if applicable
        limit_price = ""
        if zaliv_type == "Limit":
            limit_price = self.limit_price_input.text().strip()

        return {
            'token_link': token_link,
            'position_percent': position_percent,
            'zaliv_type': zaliv_type,
            'limit_price': limit_price
        }

    def validate_short_long_settings(self):
        """Validate Short/Long settings before execution"""
        settings = self.get_short_long_settings()

        # Validate token link
        if not settings['token_link']:
            QMessageBox.warning(self, "Missing Token Link", "Please enter a token link or contract address")
            return False

        # Validate position percentage
        if not settings['position_percent']:
            QMessageBox.warning(self, "Missing Position", "Please enter a position percentage")
            return False

        try:
            position = float(settings['position_percent'])
            if position <= 0 or position > 100:
                QMessageBox.warning(self, "Invalid Position", "Position percentage must be between 0 and 100")
                return False
        except ValueError:
            QMessageBox.warning(self, "Invalid Position", "Position percentage must be a valid number")
            return False

        # Validate limit price if Limit type
        if settings['zaliv_type'] == "Limit":
            if not settings['limit_price']:
                QMessageBox.warning(self, "Missing Limit Price", "Please enter a limit price")
                return False

            try:
                price = float(settings['limit_price'])
                if price <= 0:
                    QMessageBox.warning(self, "Invalid Limit Price", "Limit price must be greater than 0")
                    return False
            except ValueError:
                QMessageBox.warning(self, "Invalid Limit Price", "Limit price must be a valid number")
                return False

        return True

    def create_logs_section(self):
        """Create logs section"""
        section = QWidget()
        layout = QVBoxLayout(section)
        layout.setSpacing(10)

        # Header with title and clear button
        header_layout = QHBoxLayout()

        title = QLabel("Logs")
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        clear_btn = QPushButton("Clear Log")
        clear_btn.setObjectName("clearLogBtn")
        clear_btn.clicked.connect(self.clear_log)
        clear_btn.setFixedWidth(100)
        header_layout.addWidget(clear_btn)

        layout.addLayout(header_layout)

        # Log text area
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setMinimumHeight(150)
        self.log_output.setMaximumHeight(250)

        layout.addWidget(self.log_output)

        # Initial log message
        self.log("✓ Hysk Mexc Futures ready - You can now manage profiles")

        return section

    def refresh_profiles_table(self):
        """Refresh profiles table with data from profile manager"""
        # Stop global TOTP timer during refresh
        self.global_totp_timer.stop()
        self.totp_data.clear()

        # Disable updates during bulk operation (performance)
        self.profiles_table.setUpdatesEnabled(False)

        # Clear table
        self.profiles_table.setRowCount(0)

        # Get all profiles
        profiles = self.profile_manager.get_all_profiles()

        for profile_name in profiles:
            profile_info = self.profile_manager.get_profile_info(profile_name)
            if not profile_info:
                continue

            row = self.profiles_table.rowCount()
            self.profiles_table.insertRow(row)

            # Set row height (30% bigger than default ~30px = ~40px)
            self.profiles_table.setRowHeight(row, 45)

            # Column 0: Select checkbox
            checkbox = QCheckBox()
            checkbox.setText("")
            checkbox.setStyleSheet("""
                QCheckBox {
                    background: none;
                    border: none;
                    padding: 0px;
                    margin: 0px;
                    margin-left: 7px;
                }
                QCheckBox::indicator {
                    width: 12px;
                    height: 12px;
                    border: 2px solid #1976d2;
                    border-radius: 6px;
                    background: none;
                }
                QCheckBox::indicator:unchecked {
                    background: none;
                }
                QCheckBox::indicator:checked {
                    background-color: #1976d2;
                }
                QCheckBox::indicator:hover {
                    border-color: #64b5f6;
                }
            """)

            self.profiles_table.setCellWidget(row, 0, checkbox)

            # Column 1: Email
            email = profile_info.get('email', 'N/A')
            email_item = QTableWidgetItem(email)
            email_item.setFlags(email_item.flags() & ~Qt.ItemIsEditable)
            self.profiles_table.setItem(row, 1, email_item)

            # Column 2: Proxy (extract IP only)
            proxy = profile_info.get('proxy', '')
            proxy_ip = self.extract_proxy_ip(proxy)
            proxy_item = QTableWidgetItem(proxy_ip)
            proxy_item.setFlags(proxy_item.flags() & ~Qt.ItemIsEditable)
            self.profiles_table.setItem(row, 2, proxy_item)

            # Column 3: 2FA Code (with countdown)
            twofa_secret = profile_info.get('twofa_secret', '')
            if twofa_secret:
                # Generate initial TOTP code
                totp_text = self.generate_totp_with_timer(twofa_secret)
                totp_item = QTableWidgetItem(totp_text)
                totp_item.setForeground(QColor("#64b5f6"))
                totp_item.setFlags(totp_item.flags() & ~Qt.ItemIsEditable)
                self.profiles_table.setItem(row, 3, totp_item)

                # Store TOTP secret for this row (global timer will update)
                self.totp_data[row] = twofa_secret
            else:
                totp_item = QTableWidgetItem("—")
                totp_item.setFlags(totp_item.flags() & ~Qt.ItemIsEditable)
                self.profiles_table.setItem(row, 3, totp_item)

            # Column 4: Status
            status_item = QTableWidgetItem("Closed")
            status_item.setForeground(QColor("#ff9800"))
            status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
            self.profiles_table.setItem(row, 4, status_item)

            # Connect checkbox to update delete button visibility
            checkbox.stateChanged.connect(self.update_delete_button_visibility)

        # Re-enable updates and refresh display
        self.profiles_table.setUpdatesEnabled(True)

        # Start global TOTP timer if we have any TOTP codes to update
        if self.totp_data:
            self.global_totp_timer.start(1000)  # Update every second

    def update_all_totp_codes(self):
        """Update all TOTP codes at once (called by global timer)"""
        # Batch update all TOTP codes in one go (performance)
        for row, secret in self.totp_data.items():
            try:
                totp_text = self.generate_totp_with_timer(secret)
                item = self.profiles_table.item(row, 3)
                if item:
                    item.setText(totp_text)
            except:
                pass

    def extract_proxy_ip(self, proxy_string):
        """Extract IP address from proxy string"""
        if not proxy_string:
            return "—"

        # Format: socks5://user:pass@IP:PORT
        # Extract just the IP part
        try:
            if '@' in proxy_string:
                # Format: protocol://user:pass@IP:port
                ip_port = proxy_string.split('@')[1]
                ip = ip_port.split(':')[0]
                return ip
            else:
                return "Invalid proxy format"
        except:
            return proxy_string

    def generate_totp_with_timer(self, secret):
        """Generate TOTP code with remaining time"""
        try:
            totp = pyotp.TOTP(secret)
            code = totp.now()

            # Calculate remaining time
            import time
            current_time = int(time.time())
            time_remaining = 30 - (current_time % 30)

            return f"{code} ({time_remaining}s)"
        except:
            return "Invalid secret"

    def get_selected_profile_rows(self):
        """Get list of selected profile rows"""
        selected_rows = []
        for row in range(self.profiles_table.rowCount()):
            checkbox = self.profiles_table.cellWidget(row, 0)
            if checkbox and isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                selected_rows.append(row)
        return selected_rows

    def update_delete_button_visibility(self):
        """Show/hide delete button based on selection"""
        selected_rows = self.get_selected_profile_rows()
        if selected_rows:
            self.delete_selected_btn.show()
        else:
            self.delete_selected_btn.hide()

    def get_selected_operation_mode(self):
        """Get the currently selected operation mode"""
        checked_button = self.operation_group.checkedButton()
        if checked_button:
            return checked_button.property("mode")
        return None

    def delete_selected_profiles(self):
        """Delete all selected profiles"""
        selected_rows = self.get_selected_profile_rows()

        if not selected_rows:
            return

        # Get profile names for selected rows
        profiles_to_delete = []
        for row in selected_rows:
            email_item = self.profiles_table.item(row, 1)
            if email_item:
                email = email_item.text()

                # Find profile name by email
                profiles = self.profile_manager.get_all_profiles()
                for name in profiles:
                    info = self.profile_manager.get_profile_info(name)
                    if info and info.get('email') == email:
                        profiles_to_delete.append((name, email))
                        break

        if not profiles_to_delete:
            return

        # Confirm deletion
        profile_list = "\n".join([f"• {email}" for name, email in profiles_to_delete])
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete {len(profiles_to_delete)} selected profile(s)?\n\n{profile_list}\n\nThis cannot be undone!",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Delete all profiles
            deleted_count = 0
            for profile_name, email in profiles_to_delete:
                success, message = self.profile_manager.delete_profile(profile_name)
                if success:
                    self.log(f"🗑️ Deleted profile: {email}")
                    deleted_count += 1
                else:
                    self.log(f"❌ Failed to delete {email}: {message}")

            # Refresh table
            self.refresh_profiles_table()
            self.log(f"✅ Deleted {deleted_count} of {len(profiles_to_delete)} profile(s)")

            # Hide delete button
            self.delete_selected_btn.hide()

    # === BUTTON HANDLERS ===

    def open_settings(self):
        """Open settings dialog"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QGroupBox

        dialog = QDialog(self)
        dialog.setWindowTitle("Settings")
        dialog.setMinimumWidth(400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #0a1929;
            }
            QGroupBox {
                color: #90caf9;
                border: 1px solid #1976d2;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                background-color: #1976d2;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QLabel {
                color: #90caf9;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)

        # Excel Tools group
        excel_group = QGroupBox("Excel Tools")
        excel_layout = QVBoxLayout(excel_group)

        # Generate sample button
        generate_btn = QPushButton("📊 Generate Excel Sample")
        generate_btn.setToolTip("Create a sample Excel file with example profiles")
        generate_btn.clicked.connect(lambda: self.generate_sample_excel(dialog))
        excel_layout.addWidget(generate_btn)

        # Description
        desc_label = QLabel("Generate a sample Excel file with example data\nto use as a template for importing profiles.")
        desc_label.setStyleSheet("color: #64b5f6; font-size: 11px;")
        excel_layout.addWidget(desc_label)

        layout.addWidget(excel_group)

        # Close button
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)

        dialog.exec()

    def generate_sample_excel(self, parent_dialog=None):
        """Generate a sample Excel file with example profiles"""
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Get save location
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", "profiles_sample.xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            parent_dialog or self,
            "Save Sample Excel File",
            default_path,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Profiles"

            # Headers
            headers = ["email", "password", "proxy", "2fa_secret"]
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Sample data
            samples = [
                ["user1@example.com", "password123", "http://proxy1.example.com:8080", "JBSWY3DPEHPK3PXP"],
                ["user2@example.com", "securepass456", "socks5://proxy2.example.com:1080", "GEZDGNBVGY3TQOJQ"],
                ["user3@example.com", "mypassword789", "http://user:pass@proxy3.example.com:3128", "MFRGGZDFMY4TQMZU"],
                ["user4@example.com", "testpass000", "", ""],  # No proxy, no 2FA
                ["user5@example.com", "demo12345", "123.45.67.89:8080", "KZXW6YTBOI5HS2TN"],
            ]

            for row_num, row_data in enumerate(samples, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 20

            # Save file
            wb.save(file_path)

            self.log(f"✅ Sample Excel file generated: {file_path}")
            QMessageBox.information(
                parent_dialog or self,
                "Success",
                f"Sample Excel file created:\n{file_path}\n\nYou can use this as a template for importing profiles."
            )

        except Exception as e:
            self.log(f"❌ Error generating sample: {str(e)}")
            QMessageBox.critical(
                parent_dialog or self,
                "Error",
                f"Failed to generate sample file:\n{str(e)}"
            )

    def import_profiles(self):
        """Import profiles from Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )

        if not file_path:
            return

        reply = QMessageBox.question(
            self,
            "Confirm Import",
            f"Import profiles from:\n{file_path}\n\nExpected format:\nRow 1: email | password | proxy | 2fa_secret\nRow 2+: data\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        self.log("📥 Starting Excel import...")

        # Allow UI to update
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

        try:
            success_count, skipped_count, errors = self.profile_manager.import_from_excel(file_path)

            # Allow UI to update after import
            QApplication.processEvents()

            self.log("🔄 Refreshing profiles table...")
            QApplication.processEvents()

            self.refresh_profiles_table()

            # Allow UI to update after refresh
            QApplication.processEvents()

            result_msg = f"Import Complete!\n\n✅ Successfully imported: {success_count} profiles\n⚠️ Skipped: {skipped_count} profiles"

            if errors:
                result_msg += "\n\nErrors/Warnings:\n"
                for error in errors[:5]:
                    result_msg += f"• {error}\n"
                if len(errors) > 5:
                    result_msg += f"\n... and {len(errors) - 5} more"

            if success_count > 0:
                QMessageBox.information(self, "Import Complete", result_msg)
                self.log(f"✅ Imported {success_count} profiles successfully")
            else:
                QMessageBox.warning(self, "Import Failed", result_msg)
                self.log("❌ No profiles were imported")

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import profiles:\n{str(e)}")
            self.log(f"❌ Import error: {str(e)}")

    def open_selected_profiles(self):
        """Open selected profiles based on operation mode"""
        selected_rows = self.get_selected_profile_rows()

        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select at least one profile")
            return

        # Get selected operation mode
        operation_mode = self.get_selected_operation_mode()

        if not operation_mode:
            QMessageBox.warning(self, "No Mode Selected", "Please select an operation mode")
            return

        self.log(f"🚀 Starting operation: {operation_mode.upper()}")
        self.log(f"📊 Selected {len(selected_rows)} profile(s)")

        # Handle different operation modes
        if operation_mode == "manual":
            self.run_manual_browser_for_selected(selected_rows)
        elif operation_mode == "login":
            self.run_mexc_login_for_selected(selected_rows)
        elif operation_mode == "short":
            # Validate settings first
            if not self.validate_short_long_settings():
                return
            self.run_mexc_short_for_selected(selected_rows)
        elif operation_mode == "long":
            # Validate settings first
            if not self.validate_short_long_settings():
                return
            self.run_mexc_long_for_selected(selected_rows)
        elif operation_mode in ["balance", "rk"]:
            self.log(f"⚠️ Mode '{operation_mode.upper()}' not implemented yet")
            QMessageBox.information(
                self,
                "Coming Soon",
                f"'{operation_mode.upper()}' mode will be implemented in future version"
            )
        else:
            self.log(f"❌ Unknown mode: {operation_mode}")

    def run_manual_browser_for_selected(self, selected_rows):
        """Open browsers asynchronously using threads (non-blocking with exec())"""
        from PySide6.QtWidgets import QApplication

        self.log("🖱️ Opening browser(s) in Manual mode (async)...")
        QApplication.processEvents()

        started_count = 0

        for row in selected_rows:
            email = "Unknown"
            try:
                email_item = self.profiles_table.item(row, 1)
                if not email_item:
                    continue

                email = email_item.text()

                # Find profile by email
                profile_name = None
                profiles = self.profile_manager.get_all_profiles()
                for name in profiles:
                    info = self.profile_manager.get_profile_info(name)
                    if info and info.get('email') == email:
                        profile_name = name
                        break

                if not profile_name:
                    self.log(f"❌ Profile not found: {email}")
                    continue

                # Update status to "Opening..."
                self.update_profile_status(row, "Opening...", "#2196f3")
                QApplication.processEvents()

                # Create thread
                thread = ManualBrowserThread(
                    self.scraper_runner,
                    profile_name,
                    email,
                    headless=False
                )
                thread.row = row  # Pass row info

                # Store thread reference
                self.active_browser_threads[profile_name] = {
                    'thread': thread,
                    'email': email,
                    'row': row
                }

                # Connect signals
                thread.log_signal.connect(self.log)
                thread.driver_ready.connect(self.on_driver_ready)
                thread.error_signal.connect(self.on_browser_error)

                # Start thread (non-blocking!)
                thread.start()
                started_count += 1
                self.log(f"▶️ Started browser thread for: {email}")

                QApplication.processEvents()

            except Exception as e:
                import traceback
                self.log(f"❌ Error starting browser for {email}: {str(e)}")
                self.log(traceback.format_exc())
                QApplication.processEvents()

        self.log(f"✅ Started {started_count} browser thread(s) - UI remains responsive!")

    def on_driver_ready(self, driver, profile_info):
        """Receive Driver from thread and store it"""
        profile_name = profile_info['profile_name']
        email = profile_info['email']
        row = profile_info['row']

        # Get thread reference
        thread_info = self.active_browser_threads.get(profile_name)
        if not thread_info:
            self.log(f"⚠️ Thread not found for: {profile_name}")
            return

        # Store Driver in main thread (strong reference prevents garbage collection)
        self.active_drivers[profile_name] = {
            'driver': driver,
            'thread': thread_info['thread'],
            'email': email,
            'row': row
        }

        # Update UI
        self.update_profile_status(row, "Open (Manual)", "#4caf50")
        self.log(f"✅ Browser opened successfully: {email}")

    def on_browser_error(self, email, error_msg):
        """Handle browser creation error"""
        self.log(f"❌ Browser error for {email}")
        self.log(f"   Error details: {error_msg}")

        # Find row and update status
        for profile_name, thread_info in list(self.active_browser_threads.items()):
            if thread_info['email'] == email:
                row = thread_info['row']
                self.update_profile_status(row, "Failed to open", "#f44336")

                # Cleanup thread
                try:
                    thread = thread_info['thread']
                    thread.quit()  # Stop event loop
                    thread.wait(2000)
                except:
                    pass

                del self.active_browser_threads[profile_name]
                break

    def run_mexc_login_for_selected(self, selected_rows):
        """Run MEXC login automation for selected profiles - PARALLEL execution"""
        from PySide6.QtWidgets import QApplication
        self.log("🔐 Starting MEXC Login automation (PARALLEL mode)...")
        QApplication.processEvents()

        started_count = 0

        # Launch all threads in parallel
        for row in selected_rows:
            email_item = self.profiles_table.item(row, 1)
            if not email_item:
                continue

            email = email_item.text()

            # Find profile by email
            profile_name = None
            profiles = self.profile_manager.get_all_profiles()
            for name in profiles:
                info = self.profile_manager.get_profile_info(name)
                if info and info.get('email') == email:
                    profile_name = name
                    break

            if not profile_name:
                self.log(f"❌ Profile not found for email: {email}")
                continue

            # Get profile info
            profile_info = self.profile_manager.get_profile_info(profile_name)
            if not profile_info:
                self.log(f"❌ Failed to get profile info for: {email}")
                continue

            # Validate required fields
            password = profile_info.get('password', '')
            twofa_secret = profile_info.get('twofa_secret', '')

            if not password:
                self.log(f"❌ Missing password for: {email}")
                self.update_profile_status(row, "Error: No password", "#f44336")
                continue

            if not twofa_secret:
                self.log(f"⚠️ No 2FA secret for: {email} (will skip 2FA step)")

            # Update status to "Logging in..."
            self.update_profile_status(row, "Logging in...", "#2196f3")
            self.log(f"▶️ Starting login thread for: {email}")
            QApplication.processEvents()

            # Create and start MEXC Login thread (PARALLEL!)
            thread = MexcLoginThread(
                self.scraper_runner,
                profile_name,
                email,
                password,
                twofa_secret,
                headless=False
            )

            # Store thread reference
            self.active_threads[profile_name] = {
                'thread': thread,
                'row': row,
                'email': email
            }

            # Connect signals
            thread.log_signal.connect(self.log)
            thread.finished.connect(lambda success, result, pn=profile_name: self.on_mexc_login_finished(success, result, pn))

            # Start thread immediately (don't wait for others!)
            thread.start()
            started_count += 1

            QApplication.processEvents()

        if started_count > 0:
            self.log(f"🚀 Started {started_count} login thread(s) in PARALLEL - all running simultaneously!")
        else:
            self.log("❌ No valid profiles to process")

    def on_mexc_login_finished(self, success, result, profile_name):
        """Handle MEXC login completion"""
        thread_info = self.active_threads.get(profile_name)
        if not thread_info:
            return

        row = thread_info['row']
        email = thread_info['email']
        thread = thread_info['thread']

        if success:
            self.log(f"✅ Login successful for: {email}")
            self.update_profile_status(row, "Logged in", "#4caf50")
        else:
            self.log(f"❌ Login failed for: {email}")
            self.log(f"   Error: {result}")
            self.update_profile_status(row, "Login failed", "#f44336")

        # Wait for thread to fully finish before removing
        if thread.isRunning():
            thread.wait(1000)  # Wait max 1 second

        # Remove thread from active threads
        if profile_name in self.active_threads:
            del self.active_threads[profile_name]

        # Check if all threads completed
        if not self.active_threads:
            self.log("✅ All login threads completed!")

    def run_mexc_short_for_selected(self, selected_rows):
        """Run MEXC short position automation for selected profiles - PARALLEL execution"""
        from PySide6.QtWidgets import QApplication
        self.log("📉 Starting MEXC Short position automation (PARALLEL mode)...")
        QApplication.processEvents()

        # Get trading settings
        settings = self.get_short_long_settings()
        token_link = settings['token_link']
        position_percent = settings['position_percent']
        order_type = settings['zaliv_type']  # "Market" or "Limit"
        limit_price = settings['limit_price']

        self.log(f"🔗 Token link: {token_link}")
        self.log(f"📊 Position: {position_percent}%")
        self.log(f"📋 Order type: {order_type}")
        if order_type == "Limit":
            self.log(f"💰 Limit price: {limit_price}")

        started_count = 0

        # Launch all threads in parallel
        for row in selected_rows:
            email_item = self.profiles_table.item(row, 1)
            if not email_item:
                continue

            email = email_item.text()

            # Find profile by email
            profile_name = None
            profiles = self.profile_manager.get_all_profiles()
            for name in profiles:
                info = self.profile_manager.get_profile_info(name)
                if info and info.get('email') == email:
                    profile_name = name
                    break

            if not profile_name:
                self.log(f"❌ Profile not found for email: {email}")
                continue

            # Update status to "Opening Short..."
            self.update_profile_status(row, "Opening Short...", "#2196f3")
            self.log(f"▶️ Starting short thread for: {email}")
            QApplication.processEvents()

            # Create and start MEXC Short thread (PARALLEL!)
            thread = MexcShortThread(
                self.scraper_runner,
                profile_name,
                email,
                token_link,
                position_percent,
                order_type=order_type,
                limit_price=limit_price,
                headless=False
            )

            # Store thread reference
            self.active_trade_threads[profile_name] = {
                'thread': thread,
                'row': row,
                'email': email
            }

            # Connect signals
            thread.log_signal.connect(self.log)
            thread.finished.connect(lambda success, result, pn=profile_name: self.on_mexc_short_finished(success, result, pn))

            # Start thread immediately (don't wait for others!)
            thread.start()
            started_count += 1

            QApplication.processEvents()

        if started_count > 0:
            self.log(f"🚀 Started {started_count} short thread(s) in PARALLEL - all running simultaneously!")
        else:
            self.log("❌ No valid profiles to process")

    def on_mexc_short_finished(self, success, result, profile_name):
        """Handle MEXC short position completion"""
        thread_info = self.active_trade_threads.get(profile_name)
        if not thread_info:
            return

        row = thread_info['row']
        email = thread_info['email']
        thread = thread_info['thread']

        if success:
            self.log(f"✅ Short position opened for: {email}")
            self.update_profile_status(row, "Short opened", "#4caf50")
        else:
            self.log(f"❌ Short position failed for: {email}")
            self.log(f"   Error: {result}")
            self.update_profile_status(row, "Short failed", "#f44336")

        # Wait for thread to fully finish before removing
        if thread.isRunning():
            thread.wait(1000)  # Wait max 1 second

        # Remove thread from active threads
        if profile_name in self.active_trade_threads:
            del self.active_trade_threads[profile_name]

        # Check if all threads completed
        if not self.active_trade_threads:
            self.log("✅ All short position threads completed!")

    def run_mexc_long_for_selected(self, selected_rows):
        """Run MEXC long position automation for selected profiles - PARALLEL execution"""
        from PySide6.QtWidgets import QApplication
        self.log("📈 Starting MEXC Long position automation (PARALLEL mode)...")
        QApplication.processEvents()

        # Get trading settings
        settings = self.get_short_long_settings()
        token_link = settings['token_link']
        position_percent = settings['position_percent']
        order_type = settings['zaliv_type']  # "Market" or "Limit"
        limit_price = settings['limit_price']

        self.log(f"🔗 Token link: {token_link}")
        self.log(f"📊 Position: {position_percent}%")
        self.log(f"📋 Order type: {order_type}")
        if order_type == "Limit":
            self.log(f"💰 Limit price: {limit_price}")

        started_count = 0

        # Launch all threads in parallel
        for row in selected_rows:
            email_item = self.profiles_table.item(row, 1)
            if not email_item:
                continue

            email = email_item.text()

            # Find profile by email
            profile_name = None
            profiles = self.profile_manager.get_all_profiles()
            for name in profiles:
                info = self.profile_manager.get_profile_info(name)
                if info and info.get('email') == email:
                    profile_name = name
                    break

            if not profile_name:
                self.log(f"❌ Profile not found for email: {email}")
                continue

            # Update status to "Opening Long..."
            self.update_profile_status(row, "Opening Long...", "#2196f3")
            self.log(f"▶️ Starting long thread for: {email}")
            QApplication.processEvents()

            # Create and start MEXC Long thread (PARALLEL!)
            thread = MexcLongThread(
                self.scraper_runner,
                profile_name,
                email,
                token_link,
                position_percent,
                order_type=order_type,
                limit_price=limit_price,
                headless=False
            )

            # Store thread reference
            self.active_trade_threads[profile_name] = {
                'thread': thread,
                'row': row,
                'email': email
            }

            # Connect signals
            thread.log_signal.connect(self.log)
            thread.finished.connect(lambda success, result, pn=profile_name: self.on_mexc_long_finished(success, result, pn))

            # Start thread immediately (don't wait for others!)
            thread.start()
            started_count += 1

            QApplication.processEvents()

        if started_count > 0:
            self.log(f"🚀 Started {started_count} long thread(s) in PARALLEL - all running simultaneously!")
        else:
            self.log("❌ No valid profiles to process")

    def on_mexc_long_finished(self, success, result, profile_name):
        """Handle MEXC long position completion"""
        thread_info = self.active_trade_threads.get(profile_name)
        if not thread_info:
            return

        row = thread_info['row']
        email = thread_info['email']
        thread = thread_info['thread']

        if success:
            self.log(f"✅ Long position opened for: {email}")
            self.update_profile_status(row, "Long opened", "#4caf50")
        else:
            self.log(f"❌ Long position failed for: {email}")
            self.log(f"   Error: {result}")
            self.update_profile_status(row, "Long failed", "#f44336")

        # Wait for thread to fully finish before removing
        if thread.isRunning():
            thread.wait(1000)  # Wait max 1 second

        # Remove thread from active threads
        if profile_name in self.active_trade_threads:
            del self.active_trade_threads[profile_name]

        # Check if all threads completed
        if not self.active_trade_threads:
            self.log("✅ All long position threads completed!")

    def update_profile_status(self, row, status_text, color):
        """Update profile status in table"""
        status_item = self.profiles_table.item(row, 4)
        if status_item:
            status_item.setText(status_text)
            status_item.setForeground(QColor(color))

    def close_selected_profiles(self):
        """Close selected profiles (update status to Closed)"""
        selected_rows = self.get_selected_profile_rows()

        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select at least one profile")
            return

        self.log(f"🛑 Marking {len(selected_rows)} selected profile(s) as closed...")

        # Update status for each selected profile
        for row in selected_rows:
            email_item = self.profiles_table.item(row, 1)
            if email_item:
                self.update_profile_status(row, "Closed", "#ff9800")
                self.log(f"  • Marked as closed: {email_item.text()}")

        self.log("✅ Please manually close browser windows if needed")

    def clear_log(self):
        """Clear log output"""
        self.log_output.clear()
        self.log("✓ Log cleared")

    def log(self, message):
        """Add message to log output"""
        self.log_output.append(message)
        self.log_output.verticalScrollBar().setValue(
            self.log_output.verticalScrollBar().maximum()
        )

    def closeEvent(self, event):
        """Handle application close event - clean up threads"""
        # Check if there are active login threads
        if self.active_threads:
            self.log("⚠️ Waiting for active login threads to finish...")

            # Wait for all active threads to finish (max 5 seconds each)
            for profile_name, thread_info in list(self.active_threads.items()):
                thread = thread_info['thread']
                if thread.isRunning():
                    self.log(f"⏳ Waiting for thread: {thread_info['email']}")
                    thread.wait(5000)  # Wait max 5 seconds

                    # If still running after timeout, terminate
                    if thread.isRunning():
                        self.log(f"⚠️ Force terminating thread: {thread_info['email']}")
                        thread.terminate()
                        thread.wait(1000)

        # Check if there are active trade threads
        if self.active_trade_threads:
            self.log("⚠️ Waiting for active trade threads to finish...")

            # Wait for all active threads to finish (max 5 seconds each)
            for profile_name, thread_info in list(self.active_trade_threads.items()):
                thread = thread_info['thread']
                if thread.isRunning():
                    self.log(f"⏳ Waiting for trade thread: {thread_info['email']}")
                    thread.quit()  # Exit exec() loop
                    thread.wait(5000)  # Wait max 5 seconds

                    # If still running after timeout, terminate
                    if thread.isRunning():
                        self.log(f"⚠️ Force terminating trade thread: {thread_info['email']}")
                        thread.terminate()
                        thread.wait(1000)

            self.active_trade_threads.clear()

        # Close all active browsers and stop threads
        if self.active_drivers:
            self.log("🌐 Closing active browsers and threads...")
            for profile_name, driver_info in list(self.active_drivers.items()):
                try:
                    driver = driver_info['driver']
                    thread = driver_info.get('thread')
                    email = driver_info['email']

                    self.log(f"🔒 Closing browser: {email}")

                    # Close browser
                    driver.close()

                    # Stop thread event loop (if thread is using exec())
                    if thread:
                        self.log(f"⏹️ Stopping thread for: {email}")
                        thread.quit()  # Exit exec() loop
                        thread.wait(2000)  # Wait max 2 seconds

                except Exception as e:
                    self.log(f"⚠️ Error closing browser: {str(e)}")

            # Clear the dictionaries
            self.active_drivers.clear()
            self.active_browser_threads.clear()

        # Stop global TOTP timer
        self.global_totp_timer.stop()

        self.log("👋 Application closing...")

        # Accept the close event
        event.accept()
