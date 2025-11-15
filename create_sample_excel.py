"""
Create a sample Excel file for profile import testing

This script creates an example Excel file with the correct format for importing profiles.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Profiles"

# Headers (Row 1) - with formatting
headers = ['email', 'password', 'proxy', '2fa_secret']
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill

# Sample data (starting from Row 2)
sample_data = [
    # email, password, proxy, 2fa_secret
    ['user1@example.com', 'password123', '123.45.67.89:8080', 'JBSWY3DPEHPK3PXP'],
    ['user2@gmail.com', 'securePass456', 'http://proxy.example.com:3128', 'MFRGGZDFMZTWQ2LK'],
    ['testuser@mexc.com', 'myPassword789', '', 'GEZDGNBVGY3TQOJQ'],  # No proxy
    ['another@test.com', 'pass1234', 'socks5://45.67.89.12:1080', ''],  # SOCKS5 proxy
    ['demo@domain.com', 'demopass', '', ''],  # No proxy or 2FA
    ['vip@mexc.com', 'vippass', 'http://user:pass@1.2.3.4:8080', 'MFRGGZDFMZTWQ2LK'],  # Authenticated proxy
]

# Add sample data
for row_num, data in enumerate(sample_data, start=2):
    for col_num, value in enumerate(data, start=1):
        ws.cell(row=row_num, column=col_num, value=value)

# Adjust column widths
ws.column_dimensions['A'].width = 25  # email
ws.column_dimensions['B'].width = 20  # password
ws.column_dimensions['C'].width = 25  # proxy
ws.column_dimensions['D'].width = 25  # 2fa_secret

# Save to Desktop
desktop = Path.home() / "Desktop"
output_file = desktop / "profiles_sample.xlsx"

wb.save(output_file)

print("=" * 60)
print("✅ Sample Excel file created successfully!")
print("=" * 60)
print(f"\nFile saved to: {output_file}")
print("\nFormat:")
print("  Row 1 (headers): email | password | proxy | 2fa_secret")
print("  Row 2+: your data")
print("\nThe file contains 6 sample profiles to test with:")
print("  - 3 with proxies (HTTP, SOCKS5, Authenticated)")
print("  - 3 without proxies (direct connection)")
print("\nTo use:")
print("  1. Open the app")
print("  2. Go to Profiles tab")
print("  3. Click 'Import Profiles from Excel'")
print("  4. Select this file")
print("\n" + "=" * 60)
